from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def is_formula(value: Any) -> bool:
    """True when a cell holds a formula under a data_only=False load.

    openpyxl returns a leading-'=' string for ordinary formulas, but an
    ``ArrayFormula`` / ``DataTableFormula`` object for array/data-table formulas.
    Both must resolve to their computed value, not leak the object.
    """
    if isinstance(value, str):
        return value.startswith("=")
    return type(value).__name__ in {"ArrayFormula", "DataTableFormula"}


class ValueResolver:
    """Resolve computed cell values for a workbook.

    Formula cells return their formula string under the default (data_only=False)
    load. Agents almost always want the *number*, so this resolver surfaces it:

    1. Cached values — the results Excel/LibreOffice computed and saved in the file
       (``data_only=True``). Deterministic, cheap, no dependencies. Covers any file
       that was opened and saved by a spreadsheet app.
    2. Recompute fallback — when a formula has no cached value (e.g. a file written
       programmatically by openpyxl), optionally evaluate it with the ``formulas``
       library. Best-effort; requires the ``recompute`` extra. If unavailable, the
       cell resolves to ``None`` and ``computed_value_unavailable`` is flagged.

    Literal (non-formula) cells pass through unchanged.
    """

    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self._cached_wb: Any = None
        self._recompute_map: Any = None  # None=not built, False=unavailable, dict=values
        self.warnings: set[str] = set()

    # -- cached (primary) ------------------------------------------------
    def _cached_workbook(self) -> Any:
        if self._cached_wb is None:
            self._cached_wb = load_workbook(self.path, data_only=True)
        return self._cached_wb

    def _cached_value(self, sheet: str, coord: str) -> Any:
        try:
            wb = self._cached_workbook()
            if sheet not in wb.sheetnames:
                return None
            return wb[sheet][coord].value
        except Exception:
            return None

    # -- recompute (fallback) --------------------------------------------
    def _recompute_values(self) -> Any:
        """Lazily build a normalized {(SHEET_UPPER, COORD): value} map via `formulas`.

        Returns ``False`` when the optional dependency is missing or evaluation
        fails, so callers degrade gracefully instead of raising.
        """
        if self._recompute_map is not None:
            return self._recompute_map
        try:
            import formulas  # type: ignore

            model = formulas.ExcelModel().loads(str(self.path)).finish()
            solution = model.calculate()
            resolved: dict[tuple[str, str], Any] = {}
            for key, cell in solution.items():
                # keys look like "'[FILE.XLSX]SHEET'!A1"
                if "]" not in key or "!" not in key:
                    continue
                sheet_part, coord = key.rsplit("!", 1)
                sheet_name = sheet_part.split("]", 1)[1].rstrip("'").upper()
                value = getattr(cell, "value", cell)
                try:
                    # formulas returns numpy-wrapped scalars; unwrap 1x1 arrays
                    if hasattr(value, "ravel"):
                        flat = value.ravel()
                        value = flat[0] if len(flat) else None
                except Exception:
                    pass
                resolved[(sheet_name, coord.replace("$", "").upper())] = value
            self._recompute_map = resolved
        except Exception:
            self._recompute_map = False
        return self._recompute_map

    def _recompute_value(self, sheet: str, coord: str) -> Any:
        table = self._recompute_values()
        if not table:
            return None
        return table.get((sheet.upper(), coord.replace("$", "").upper()))

    # -- public ----------------------------------------------------------
    def resolve(self, sheet: str, coord: str, raw: Any) -> Any:
        """Return the computed value for a cell.

        ``raw`` is the value from the formula workbook (formula string or literal).
        """
        if not is_formula(raw):
            return raw
        cached = self._cached_value(sheet, coord)
        if cached is not None:
            return cached
        recomputed = self._recompute_value(sheet, coord)
        if recomputed is not None:
            return recomputed
        self.warnings.add("computed_value_unavailable")
        return None


def get_value_resolver(session: Any) -> ValueResolver:
    """Lazily attach and return a ValueResolver for a workbook session."""
    resolver = getattr(session, "value_resolver", None)
    if resolver is None:
        resolver = ValueResolver(session.path)
        session.value_resolver = resolver
    return resolver
