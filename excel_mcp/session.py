from __future__ import annotations

import sqlite3
import time
import uuid
import json
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from excel_mcp.config import ExcelMcpConfig


@dataclass
class WorkbookSession:
    session_id: str
    path: Path
    file_size: int
    mtime: float
    workbook: Workbook
    opened_at: float
    last_access_at: float
    duckdb_connection: Any = None
    regions: list[Any] = field(default_factory=list)
    staged: dict[str, Any] = field(default_factory=dict)
    recent_read_ranges: list[str] = field(default_factory=list)
    value_resolver: Any = None

    @property
    def fingerprint(self) -> str:
        return f"{self.path}:{self.file_size}:{self.mtime}"


class SessionRegistry:
    def __init__(self, config: ExcelMcpConfig):
        self.config = config
        self.config.cache_dir.mkdir(parents=True, exist_ok=True)
        self._sessions: OrderedDict[str, WorkbookSession] = OrderedDict()
        self._fingerprint_to_session: dict[str, str] = {}
        self._db_path = self.config.cache_dir / "sessions.sqlite3"
        self._init_db()

    def open(self, path: Path) -> tuple[WorkbookSession, bool]:
        stat = path.stat()
        fingerprint = f"{path}:{stat.st_size}:{stat.st_mtime}"
        existing_id = self._fingerprint_to_session.get(fingerprint)
        if existing_id and existing_id in self._sessions:
            session = self._sessions[existing_id]
            session.last_access_at = time.time()
            self._sessions.move_to_end(existing_id)
            self._record_session(session)
            return session, True

        workbook = load_workbook(path, data_only=False)
        now = time.time()
        session = WorkbookSession(
            session_id=f"ses_{uuid.uuid4().hex[:16]}",
            path=path,
            file_size=stat.st_size,
            mtime=stat.st_mtime,
            workbook=workbook,
            opened_at=now,
            last_access_at=now,
        )
        self._sessions[session.session_id] = session
        self._fingerprint_to_session[fingerprint] = session.session_id
        self._record_session(session)
        self._evict_if_needed()
        return session, False

    def get(self, session_id: str) -> WorkbookSession:
        session = self._sessions.get(session_id)
        if session is None:
            session = self._load_session(session_id)
            if session is None:
                raise KeyError(session_id)
            self._sessions[session.session_id] = session
            self._fingerprint_to_session[session.fingerprint] = session.session_id
            self._evict_if_needed()
        stat = session.path.stat()
        if stat.st_size != session.file_size or stat.st_mtime != session.mtime:
            self.close(session_id)
            raise KeyError(session_id)
        session.last_access_at = time.time()
        self._load_staged(session)
        self._sessions.move_to_end(session_id)
        self._record_session(session)
        return session

    def close(self, session_id: str) -> None:
        session = self._sessions.pop(session_id, None)
        if session is None:
            return
        self._fingerprint_to_session.pop(session.fingerprint, None)
        if session.duckdb_connection is not None:
            try:
                session.duckdb_connection.close()
            except Exception:
                pass

    def _evict_if_needed(self) -> None:
        while len(self._sessions) > self.config.max_open_workbooks:
            session_id = next(iter(self._sessions))
            self.close(session_id)

    def _init_db(self) -> None:
        with sqlite3.connect(self._db_path) as conn:
            conn.execute(
                """
                create table if not exists sessions (
                    session_id text primary key,
                    path text not null,
                    file_size integer not null,
                    mtime real not null,
                    opened_at real not null,
                    last_access_at real not null
                )
                """
            )
            conn.execute(
                """
                create table if not exists staged_writes (
                    staged_id text primary key,
                    session_id text not null,
                    operations_json text not null,
                    touched_ranges_json text not null,
                    changes_json text not null,
                    warnings_json text not null,
                    created_at real not null
                )
                """
            )
            conn.execute(
                """
                create table if not exists audit_events (
                    event_id text primary key,
                    created_at real not null,
                    operation text not null,
                    session_id text,
                    path text,
                    input_summary_json text not null,
                    source_refs_json text not null,
                    touched_ranges_json text not null,
                    warnings_json text not null,
                    ok integer not null,
                    error_code text,
                    elapsed_ms integer
                )
                """
            )

    def _record_session(self, session: WorkbookSession) -> None:
        with sqlite3.connect(self._db_path) as conn:
            conn.execute(
                """
                insert or replace into sessions
                (session_id, path, file_size, mtime, opened_at, last_access_at)
                values (?, ?, ?, ?, ?, ?)
                """,
                (
                    session.session_id,
                    str(session.path),
                    session.file_size,
                    session.mtime,
                    session.opened_at,
                    session.last_access_at,
                ),
            )

    def save_staged_write(
        self,
        session_id: str,
        staged_id: str,
        operations: list[dict],
        touched_ranges: list[str],
        changes: list[dict],
        warnings: list[str],
    ) -> None:
        with sqlite3.connect(self._db_path) as conn:
            conn.execute(
                """
                insert or replace into staged_writes
                (staged_id, session_id, operations_json, touched_ranges_json, changes_json, warnings_json, created_at)
                values (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    staged_id,
                    session_id,
                    json.dumps(operations, ensure_ascii=False),
                    json.dumps(touched_ranges, ensure_ascii=False),
                    json.dumps(changes, ensure_ascii=False),
                    json.dumps(warnings, ensure_ascii=False),
                    time.time(),
                ),
            )

    def record_audit_event(
        self,
        operation: str,
        session_id: str | None,
        path: str | None,
        input_summary: dict,
        source_refs: list[str],
        touched_ranges: list[str],
        warnings: list[str],
        ok: bool,
        error_code: str | None,
        elapsed_ms: int | None,
    ) -> str:
        event_id = f"evt_{uuid.uuid4().hex[:16]}"
        with sqlite3.connect(self._db_path) as conn:
            conn.execute(
                """
                insert into audit_events
                (event_id, created_at, operation, session_id, path, input_summary_json,
                 source_refs_json, touched_ranges_json, warnings_json, ok, error_code, elapsed_ms)
                values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_id,
                    time.time(),
                    operation,
                    session_id,
                    path,
                    json.dumps(input_summary, ensure_ascii=False),
                    json.dumps(source_refs, ensure_ascii=False),
                    json.dumps(touched_ranges, ensure_ascii=False),
                    json.dumps(warnings, ensure_ascii=False),
                    1 if ok else 0,
                    error_code,
                    elapsed_ms,
                ),
            )
        return event_id

    def list_audit_events(
        self,
        session_id: str | None = None,
        path: str | None = None,
        limit: int = 100,
    ) -> list[dict]:
        clauses = []
        params: list[object] = []
        if session_id:
            clauses.append("session_id = ?")
            params.append(session_id)
        if path:
            clauses.append("path = ?")
            params.append(path)
        where = f"where {' and '.join(clauses)}" if clauses else ""
        params.append(limit)
        with sqlite3.connect(self._db_path) as conn:
            rows = conn.execute(
                f"""
                select event_id, created_at, operation, session_id, path, input_summary_json,
                       source_refs_json, touched_ranges_json, warnings_json, ok, error_code, elapsed_ms
                from audit_events
                {where}
                order by created_at desc
                limit ?
                """,
                params,
            ).fetchall()
        return [
            {
                "event_id": row[0],
                "created_at": row[1],
                "operation": row[2],
                "session_id": row[3],
                "path": row[4],
                "input_summary": json.loads(row[5]),
                "source_refs": json.loads(row[6]),
                "touched_ranges": json.loads(row[7]),
                "warnings": json.loads(row[8]),
                "ok": bool(row[9]),
                "error_code": row[10],
                "elapsed_ms": row[11],
            }
            for row in rows
        ]

    def _load_session(self, session_id: str) -> WorkbookSession | None:
        with sqlite3.connect(self._db_path) as conn:
            row = conn.execute(
                """
                select session_id, path, file_size, mtime, opened_at, last_access_at
                from sessions
                where session_id = ?
                """,
                (session_id,),
            ).fetchone()
        if row is None:
            return None
        path = Path(row[1])
        if not path.exists():
            return None
        workbook = load_workbook(path, data_only=False)
        return WorkbookSession(
            session_id=row[0],
            path=path,
            file_size=row[2],
            mtime=row[3],
            workbook=workbook,
            opened_at=row[4],
            last_access_at=row[5],
        )

    def _load_staged(self, session: WorkbookSession) -> None:
        with sqlite3.connect(self._db_path) as conn:
            rows = conn.execute(
                """
                select staged_id, operations_json, touched_ranges_json, changes_json, warnings_json
                from staged_writes
                where session_id = ?
                """,
                (session.session_id,),
            ).fetchall()
        for row in rows:
            session.staged[row[0]] = {
                "staged_id": row[0],
                "operations": json.loads(row[1]),
                "touched_ranges": json.loads(row[2]),
                "changes": json.loads(row[3]),
                "warnings": json.loads(row[4]),
            }
