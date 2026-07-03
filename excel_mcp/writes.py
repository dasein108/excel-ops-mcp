from __future__ import annotations

import copy
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from excel_mcp.config import ExcelMcpConfig
from excel_mcp.policy import PathPolicy, PolicyError
from excel_mcp.schemas import (
    SpreadsheetCommitRequest,
    SpreadsheetCommitResponse,
    SpreadsheetDiffResponse,
    SpreadsheetWriteRequest,
    SpreadsheetWriteResponse,
)
from excel_mcp.session import WorkbookSession
from excel_mcp.utils import bounds_to_a1, cell_value_for_json, source_ref


@dataclass
class StagedWrite:
    staged_id: str
    operations: list[dict[str, Any]]
    touched_ranges: list[str]
    changes: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)


def stage_write(session: WorkbookSession, request: SpreadsheetWriteRequest) -> SpreadsheetWriteResponse:
    rejected: list[dict[str, Any]] = []
    touched_ranges: list[str] = []
    changes: list[dict[str, Any]] = []
    warnings: list[str] = []

    for index, operation in enumerate(request.operations):
        try:
            op_changes, op_ranges, op_warnings = preview_operation(session, operation)
            changes.extend(op_changes)
            touched_ranges.extend(op_ranges)
            warnings.extend(op_warnings)
        except PolicyError as exc:
            rejected.append({"index": index, "code": exc.code, "message": exc.message, "operation": operation})
        except Exception as exc:
            rejected.append({"index": index, "code": "invalid_operation", "message": str(exc), "operation": operation})

    accepted_count = len(request.operations) - len(rejected)
    staged_id = None
    if accepted_count:
        staged_id = f"stg_{uuid.uuid4().hex[:16]}"
        accepted_ops = [op for idx, op in enumerate(request.operations) if not any(item["index"] == idx for item in rejected)]
        session.staged[staged_id] = StagedWrite(
            staged_id=staged_id,
            operations=accepted_ops,
            touched_ranges=_dedupe(touched_ranges),
            changes=changes,
            warnings=_dedupe(warnings),
        )

    return SpreadsheetWriteResponse(
        ok=not rejected,
        session_id=session.session_id,
        staged_id=staged_id,
        accepted_operations=accepted_count,
        rejected_operations=rejected,
        touched_ranges=_dedupe(touched_ranges),
        changes=changes,
        warnings=_dedupe(warnings),
        source_refs=[source_ref(session.path, sheet, range_text) for sheet, range_text in _split_touched_refs(touched_ranges)],
    )


def diff_staged(session: WorkbookSession, staged_id: str | None = None) -> SpreadsheetDiffResponse:
    staged = _get_staged(session, staged_id)
    return SpreadsheetDiffResponse(
        ok=True,
        session_id=session.session_id,
        changed_ranges=staged.touched_ranges,
        changes=staged.changes,
        warnings=staged.warnings,
        source_refs=[source_ref(session.path, sheet, range_text) for sheet, range_text in _split_touched_refs(staged.touched_ranges)],
    )


def commit_staged(session: WorkbookSession, request: SpreadsheetCommitRequest, config: ExcelMcpConfig) -> SpreadsheetCommitResponse:
    staged = _get_staged(session, request.staged_id)
    output_path = _resolve_output_path(session, request, config)

    workbook = load_workbook(session.path, data_only=False)
    for operation in staged.operations:
        apply_operation(workbook[operation["sheet"]], operation)
    workbook.save(output_path)

    return SpreadsheetCommitResponse(
        ok=True,
        session_id=session.session_id,
        output_path=str(output_path),
        changed_ranges=staged.touched_ranges,
        changes=staged.changes,
        warnings=staged.warnings,
        source_refs=[source_ref(session.path, sheet, range_text) for sheet, range_text in _split_touched_refs(staged.touched_ranges)],
    )


def preview_operation(session: WorkbookSession, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    op_type = operation.get("type")
    sheet_name = operation.get("sheet")
    if not sheet_name or sheet_name not in session.workbook.sheetnames:
        raise PolicyError("sheet_not_found", "Operation sheet does not exist.")
    ws = session.workbook[sheet_name]

    if op_type == "set_values":
        return _preview_set_values(ws, operation)
    if op_type == "set_formula":
        return _preview_set_formula(ws, operation)
    if op_type == "clear_range":
        return _preview_clear_range(ws, operation)
    if op_type == "append_rows":
        return _preview_append_rows(ws, operation)
    if op_type == "insert_rows":
        return _preview_insert_rows(ws, operation)
    if op_type == "delete_rows":
        return _preview_delete_rows(ws, operation)
    if op_type == "copy_range":
        return _preview_copy_range(ws, operation)
    raise PolicyError("unsupported_operation", f"Unsupported write operation: {op_type}")


def apply_operation(ws: Worksheet, operation: dict[str, Any]) -> None:
    op_type = operation["type"]
    if op_type == "set_values":
        min_col, min_row, _, _ = range_boundaries(_single_cell_range(operation["start"]))
        for row_offset, row in enumerate(operation.get("values", [])):
            for col_offset, value in enumerate(row):
                ws.cell(min_row + row_offset, min_col + col_offset).value = value
        return
    if op_type == "set_formula":
        ws[operation["cell"]].value = operation["formula"]
        return
    if op_type == "clear_range":
        min_col, min_row, max_col, max_row = range_boundaries(operation["range"])
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).value = None
        return
    if op_type == "append_rows":
        for row in operation.get("rows", []):
            ws.append(row)
        return
    if op_type == "insert_rows":
        ws.insert_rows(int(operation["idx"]), int(operation.get("amount", 1)))
        return
    if op_type == "delete_rows":
        ws.delete_rows(int(operation["idx"]), int(operation.get("amount", 1)))
        return
    if op_type == "copy_range":
        _apply_copy_range(ws, operation)
        return
    raise PolicyError("unsupported_operation", f"Unsupported write operation: {op_type}")


def _preview_set_values(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    values = operation.get("values")
    if not isinstance(values, list) or not values or not all(isinstance(row, list) for row in values):
        raise PolicyError("invalid_values", "set_values requires a non-empty 2D values array.")
    min_col, min_row, _, _ = range_boundaries(_single_cell_range(operation.get("start")))
    max_row = min_row + len(values) - 1
    max_col = min_col + max(len(row) for row in values) - 1
    bounds = bounds_to_a1(min_row, min_col, max_row, max_col)
    changes = []
    for row_offset, row in enumerate(values):
        for col_offset, value in enumerate(row):
            cell = ws.cell(min_row + row_offset, min_col + col_offset)
            changes.append(_change(ws.title, cell.coordinate, cell.value, value))
    return changes, [f"{ws.title}!{bounds}"], _warnings_for_range(ws, bounds)


def _preview_set_formula(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    cell_ref = operation.get("cell")
    formula = operation.get("formula")
    if not isinstance(cell_ref, str) or not isinstance(formula, str) or not formula.startswith("="):
        raise PolicyError("invalid_formula", "set_formula requires a cell and a formula starting with '='.")
    cell = ws[cell_ref]
    return [_change(ws.title, cell.coordinate, cell.value, formula)], [f"{ws.title}!{cell.coordinate}:{cell.coordinate}"], _warnings_for_range(ws, f"{cell.coordinate}:{cell.coordinate}")


def _preview_clear_range(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    range_text = operation.get("range")
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    changes = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                changes.append(_change(ws.title, cell.coordinate, cell.value, None))
    bounds = bounds_to_a1(min_row, min_col, max_row, max_col)
    return changes, [f"{ws.title}!{bounds}"], _warnings_for_range(ws, bounds)


def _preview_append_rows(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    rows = operation.get("rows")
    if not isinstance(rows, list) or not rows or not all(isinstance(row, list) for row in rows):
        raise PolicyError("invalid_rows", "append_rows requires a non-empty 2D rows array.")
    start_row = ws.max_row + 1
    max_cols = max(len(row) for row in rows)
    bounds = bounds_to_a1(start_row, 1, start_row + len(rows) - 1, max_cols)
    changes = []
    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):
            changes.append(_change(ws.title, ws.cell(start_row + row_offset, 1 + col_offset).coordinate, None, value))
    return changes, [f"{ws.title}!{bounds}"], []


def _preview_insert_rows(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    idx = int(operation.get("idx", 0))
    amount = int(operation.get("amount", 1))
    if idx < 1 or amount < 1:
        raise PolicyError("invalid_row_operation", "insert_rows requires idx>=1 and amount>=1.")
    bounds = bounds_to_a1(idx, 1, idx + amount - 1, ws.max_column)
    return [{"sheet": ws.title, "range": bounds, "before": None, "after": f"{amount} inserted row(s)"}], [f"{ws.title}!{bounds}"], []


def _preview_delete_rows(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    idx = int(operation.get("idx", 0))
    amount = int(operation.get("amount", 1))
    if idx < 1 or amount < 1:
        raise PolicyError("invalid_row_operation", "delete_rows requires idx>=1 and amount>=1.")
    bounds = bounds_to_a1(idx, 1, min(ws.max_row, idx + amount - 1), ws.max_column)
    return [{"sheet": ws.title, "range": bounds, "before": f"{amount} row(s)", "after": None}], [f"{ws.title}!{bounds}"], _warnings_for_range(ws, bounds)


def _preview_copy_range(ws: Worksheet, operation: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    source = operation.get("source")
    target = operation.get("target")
    min_col, min_row, max_col, max_row = range_boundaries(source)
    target_col, target_row, _, _ = range_boundaries(_single_cell_range(target))
    height = max_row - min_row + 1
    width = max_col - min_col + 1
    target_bounds = bounds_to_a1(target_row, target_col, target_row + height - 1, target_col + width - 1)
    changes = []
    for row_offset in range(height):
        for col_offset in range(width):
            source_cell = ws.cell(min_row + row_offset, min_col + col_offset)
            target_cell = ws.cell(target_row + row_offset, target_col + col_offset)
            changes.append(_change(ws.title, target_cell.coordinate, target_cell.value, source_cell.value))
    return changes, [f"{ws.title}!{target_bounds}"], _warnings_for_range(ws, target_bounds)


def _apply_copy_range(ws: Worksheet, operation: dict[str, Any]) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(operation["source"])
    target_col, target_row, _, _ = range_boundaries(_single_cell_range(operation["target"]))
    for row_offset in range(max_row - min_row + 1):
        for col_offset in range(max_col - min_col + 1):
            source_cell = ws.cell(min_row + row_offset, min_col + col_offset)
            target_cell = ws.cell(target_row + row_offset, target_col + col_offset)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format


def _warnings_for_range(ws: Worksheet, bounds: str) -> list[str]:
    warnings: list[str] = []
    min_col, min_row, max_col, max_row = range_boundaries(bounds)
    formulas = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.startswith("="):
                formulas += 1
    if formulas:
        warnings.append(f"{ws.title}!{bounds} touches {formulas} formula cell(s).")
    if any(_ranges_intersect(bounds, str(merged)) for merged in ws.merged_cells.ranges):
        warnings.append(f"{ws.title}!{bounds} intersects merged cells.")
    if ws.protection.sheet:
        warnings.append(f"{ws.title} is protected.")
    return warnings


def _ranges_intersect(left: str, right: str) -> bool:
    l_min_col, l_min_row, l_max_col, l_max_row = range_boundaries(left)
    r_min_col, r_min_row, r_max_col, r_max_row = range_boundaries(right)
    return not (l_max_col < r_min_col or r_max_col < l_min_col or l_max_row < r_min_row or r_max_row < l_min_row)


def _resolve_output_path(session: WorkbookSession, request: SpreadsheetCommitRequest, config: ExcelMcpConfig) -> Path:
    if request.output_path:
        return PathPolicy(config).validate_output_file(request.output_path, overwrite=request.overwrite)
    default = session.path.with_name(f"{session.path.stem}.updated{session.path.suffix}")
    return PathPolicy(config).validate_output_file(str(default), overwrite=request.overwrite)


def _get_staged(session: WorkbookSession, staged_id: str | None) -> StagedWrite:
    if staged_id is None:
        if len(session.staged) != 1:
            raise PolicyError("staged_id_required", "staged_id is required when zero or multiple staged writes exist.")
        return next(iter(session.staged.values()))
    staged = session.staged.get(staged_id)
    if staged is None:
        raise PolicyError("staged_not_found", "Staged write does not exist.")
    if isinstance(staged, dict):
        return StagedWrite(
            staged_id=staged["staged_id"],
            operations=staged["operations"],
            touched_ranges=staged["touched_ranges"],
            changes=staged["changes"],
            warnings=staged["warnings"],
        )
    return staged


def _single_cell_range(value: Any) -> str:
    if not isinstance(value, str) or not value:
        raise PolicyError("invalid_cell", "A cell reference is required.")
    return f"{value}:{value}"


def _change(sheet: str, cell: str, before: Any, after: Any) -> dict[str, Any]:
    return {"sheet": sheet, "cell": cell, "before": cell_value_for_json(before), "after": cell_value_for_json(after)}


def _dedupe(values: list[str]) -> list[str]:
    return list(dict.fromkeys(values))


def _split_touched_refs(touched_ranges: list[str]) -> list[tuple[str, str]]:
    refs: list[tuple[str, str]] = []
    for item in touched_ranges:
        if "!" not in item:
            continue
        sheet, range_text = item.split("!", 1)
        refs.append((sheet, range_text))
    return refs
