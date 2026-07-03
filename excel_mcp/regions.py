from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from excel_mcp.schemas import ColumnInfo, RegionInfo
from excel_mcp.utils import bounds_to_a1, cell_value_for_json, dedupe_names, safe_identifier, source_ref


@dataclass
class RawRegion:
    sheet: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def bounds(self) -> str:
        return bounds_to_a1(self.min_row, self.min_col, self.max_row, self.max_col)


def detect_regions(path: Path, ws: Worksheet) -> list[RegionInfo]:
    raw_regions = _split_regions(ws)
    period_row = _sheet_period_row(ws)
    out: list[RegionInfo] = []
    for index, raw in enumerate(raw_regions, start=1):
        if _is_matrix_region(ws, raw, period_row):
            # Transposed layout: time periods run across columns, labels in column A.
            # Name columns from the sheet's period header (not the band's own row) so
            # the SQL table reads like `line_item, jan, feb, …, year_2` and every row
            # is data (no header row to skip).
            header_row = None
            columns = _matrix_columns(ws, raw, period_row, path)
            data_start = raw.min_row
            orientation = "matrix"
            row_count = max(0, raw.max_row - raw.min_row + 1)
        else:
            header_row = _infer_header_row(ws, raw)
            columns = _columns(ws, raw, header_row, path)
            data_start = header_row + 1 if header_row else raw.min_row
            orientation = "tabular"
            row_count = max(0, raw.max_row - (header_row or raw.min_row))
        sample_rows = _sample_rows(ws, raw, header_row, columns)
        kind = _classify_region(columns, raw, sample_rows)
        table_name = safe_identifier(f"{ws.title}_{kind}_{index}", f"region_{index}")
        region_id = f"{safe_identifier(ws.title, 'sheet')}_r{index}"
        out.append(
            RegionInfo(
                region_id=region_id,
                table_name=table_name,
                sheet=ws.title,
                bounds=raw.bounds,
                header_row=header_row,
                data_start_row=data_start,
                row_count=row_count,
                column_count=raw.max_col - raw.min_col + 1,
                region_kind=kind,
                orientation=orientation,
                confidence=_confidence(ws, raw, header_row),
                columns=columns,
                sample_rows=sample_rows,
                source_ref=source_ref(path, ws.title, raw.bounds),
            )
        )
    return out


def _split_regions(ws: Worksheet) -> list[RawRegion]:
    min_row, max_row = ws.min_row, ws.max_row
    min_col, max_col = ws.min_column, ws.max_column
    non_empty_rows = [r for r in range(min_row, max_row + 1) if any(_is_non_empty(ws.cell(r, c).value) for c in range(min_col, max_col + 1))]
    if not non_empty_rows:
        return []

    row_bands = _contiguous_bands(non_empty_rows)
    regions: list[RawRegion] = []
    for row_start, row_end in row_bands:
        non_empty_cols = [
            c for c in range(min_col, max_col + 1) if any(_is_non_empty(ws.cell(r, c).value) for r in range(row_start, row_end + 1))
        ]
        for col_start, col_end in _contiguous_bands(non_empty_cols):
            # Trim fully-empty edges inside the band.
            regions.append(RawRegion(ws.title, row_start, col_start, row_end, col_end))
    return regions


def _contiguous_bands(values: list[int]) -> list[tuple[int, int]]:
    if not values:
        return []
    bands: list[tuple[int, int]] = []
    start = prev = values[0]
    for value in values[1:]:
        if value == prev + 1:
            prev = value
            continue
        bands.append((start, prev))
        start = prev = value
    bands.append((start, prev))
    return bands


def _infer_header_row(ws: Worksheet, raw: RawRegion) -> int | None:
    best_row: int | None = None
    best_score = -1.0
    max_scan = min(raw.max_row, raw.min_row + 4)
    for row in range(raw.min_row, max_scan + 1):
        values = [ws.cell(row, col).value for col in range(raw.min_col, raw.max_col + 1)]
        score = _header_score(values)
        if score > best_score:
            best_row = row
            best_score = score
    if best_row is None or best_score < 0.35:
        return None
    return best_row


def _header_score(values: list[Any]) -> float:
    non_empty = [v for v in values if _is_non_empty(v)]
    if not values or not non_empty:
        return 0.0
    fill = len(non_empty) / len(values)
    strings = sum(isinstance(v, str) for v in non_empty) / len(non_empty)
    unique = len({str(v).strip().lower() for v in non_empty}) / len(non_empty)
    shortish = sum(len(str(v)) <= 60 for v in non_empty) / len(non_empty)
    return 0.3 * fill + 0.3 * strings + 0.25 * unique + 0.15 * shortish


def _columns(ws: Worksheet, raw: RawRegion, header_row: int | None, path: Path) -> list[ColumnInfo]:
    if header_row:
        names = [str(ws.cell(header_row, c).value or f"column_{c - raw.min_col + 1}") for c in range(raw.min_col, raw.max_col + 1)]
    else:
        names = [f"column_{i}" for i in range(1, raw.max_col - raw.min_col + 2)]
    deduped = dedupe_names(names)
    columns: list[ColumnInfo] = []
    for offset, name in enumerate(deduped):
        col = raw.min_col + offset
        values = [ws.cell(r, col).value for r in range((header_row + 1 if header_row else raw.min_row), raw.max_row + 1)]
        columns.append(
            ColumnInfo(
                name=name,
                index=offset + 1,
                type=_infer_type(values),
                source_ref=source_ref(path, ws.title, bounds_to_a1(raw.min_row, col, raw.max_row, col)),
            )
        )
    return columns


def _sample_rows(ws: Worksheet, raw: RawRegion, header_row: int | None, columns: list[ColumnInfo]) -> list[dict[str, Any]]:
    start = header_row + 1 if header_row else raw.min_row
    rows: list[dict[str, Any]] = []
    for row in range(start, min(raw.max_row, start + 2) + 1):
        item: dict[str, Any] = {}
        for offset, column in enumerate(columns):
            item[column.name] = cell_value_for_json(ws.cell(row, raw.min_col + offset).value)
        if any(value is not None for value in item.values()):
            rows.append(item)
    return rows


_PERIOD_TOKENS = (
    "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
    "january", "february", "march", "april", "june", "july", "august", "september",
    "october", "november", "december",
    "q1", "q2", "q3", "q4", "year", "month", "week", "fy",
)


_MONTH_TOKENS = ("jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec")


def _period_hits(values: list[Any]) -> int:
    hits = 0
    for value in values:
        token = str(value).strip().lower() if value is not None else ""
        if any(token == p or token.startswith(p) for p in _PERIOD_TOKENS):
            hits += 1
    return hits


def _sheet_period_row(ws: Worksheet) -> int | None:
    """The sheet's time-axis header row: the first row with >=3 period tokens.

    In transposed models this row (e.g. `<title> | Jan | Feb … | Year 2 …`) is the
    real header for every data band below it, even though the bands are detected as
    separate regions.
    """
    max_scan = min(ws.max_row, (ws.min_row or 1) + 6)
    for row in range(ws.min_row or 1, max_scan + 1):
        values = [ws.cell(row, c).value for c in range(ws.min_column or 1, (ws.max_column or 1) + 1)]
        if _period_hits(values) >= 3:
            return row
    return None


def _is_matrix_region(ws: Worksheet, raw: RawRegion, period_row: int | None) -> bool:
    """A region is matrix when the sheet has a period row whose tokens line up over
    this region's columns and column A carries a text label."""
    if not period_row:
        return False
    values = [ws.cell(period_row, c).value for c in range(raw.min_col, raw.max_col + 1)]
    if _period_hits(values) < 3:
        return False
    first_col_values = [ws.cell(r, raw.min_col).value for r in range(raw.min_row, raw.max_row + 1)]
    return any(isinstance(v, str) for v in first_col_values if _is_non_empty(v))


def _matrix_columns(ws: Worksheet, raw: RawRegion, period_row: int, path: Path) -> list[ColumnInfo]:
    names: list[str] = []
    for offset, c in enumerate(range(raw.min_col, raw.max_col + 1)):
        if offset == 0:
            names.append("line_item")
            continue
        token = ws.cell(period_row, c).value
        names.append(str(token) if _is_non_empty(token) else f"column_{offset + 1}")
    deduped = dedupe_names(names)
    columns: list[ColumnInfo] = []
    for offset, name in enumerate(deduped):
        col = raw.min_col + offset
        values = [ws.cell(r, col).value for r in range(raw.min_row, raw.max_row + 1)]
        columns.append(
            ColumnInfo(
                name=name,
                index=offset + 1,
                type=_infer_type(values),
                source_ref=source_ref(path, ws.title, bounds_to_a1(raw.min_row, col, raw.max_row, col)),
            )
        )
    return columns


def matrix_month_columns(column_names: list[str]) -> list[str]:
    """The subset of column names that are Year-1 months, in calendar order."""
    lowered = {name.lower(): name for name in column_names}
    return [lowered[m] for m in _MONTH_TOKENS if m in lowered]


def _classify_region(columns: list[ColumnInfo], raw: RawRegion, sample_rows: list[dict[str, Any]]) -> str:
    names = " ".join(column.name.lower() for column in columns)
    if any(term in names for term in ["amount", "debit", "credit", "transaction", "vendor", "invoice", "date"]):
        return "ledger"
    if raw.max_row - raw.min_row <= 6 and raw.max_col - raw.min_col <= 3:
        return "parameters"
    if any(term in names for term in ["total", "summary", "variance", "month"]):
        return "summary"
    if len(sample_rows) >= 1 and len(columns) >= 2:
        return "table"
    return "unknown"


def _confidence(ws: Worksheet, raw: RawRegion, header_row: int | None) -> float:
    area = (raw.max_row - raw.min_row + 1) * (raw.max_col - raw.min_col + 1)
    filled = sum(
        1
        for row in range(raw.min_row, raw.max_row + 1)
        for col in range(raw.min_col, raw.max_col + 1)
        if _is_non_empty(ws.cell(row, col).value)
    )
    density = filled / area if area else 0.0
    header_bonus = 0.2 if header_row else 0.0
    return round(min(1.0, density * 0.8 + header_bonus), 2)


def _infer_type(values: list[Any]) -> str:
    non_empty = [v for v in values if _is_non_empty(v)]
    if not non_empty:
        return "text"
    if all(isinstance(v, bool) for v in non_empty):
        return "boolean"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_empty):
        return "integer"
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_empty):
        return "decimal"
    if all(hasattr(v, "date") or hasattr(v, "isoformat") for v in non_empty):
        return "date"
    return "text"


def _is_non_empty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""

