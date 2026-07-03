from __future__ import annotations

import re
from typing import Any

from openpyxl.utils.cell import range_boundaries, get_column_letter

from excel_mcp.policy import PolicyError
from excel_mcp.schemas import SpreadsheetTraceResponse
from excel_mcp.session import WorkbookSession
from excel_mcp.utils import cell_value_for_json, source_ref
from excel_mcp.values import get_value_resolver, is_formula

# A cell/range reference with an optional sheet qualifier:
#   'P&L Projection'!B15:M15   |   Assumptions!B7   |   B11   |   A1:B2
_REF = re.compile(
    r"(?:(?:'(?P<qsheet>[^']+)'|(?P<sheet>[A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"(?P<a1>\$?[A-Z]{1,3}\$?[0-9]+(?::\$?[A-Z]{1,3}\$?[0-9]+)?)"
)

MAX_TRACE_DEPTH = 5
_RANGE_VALUE_CAP = 32


def _formula_text(raw: Any) -> str | None:
    """Formula source for an openpyxl cell value (plain string or ArrayFormula)."""
    if isinstance(raw, str) and raw.startswith("="):
        return raw
    text = getattr(raw, "text", None)
    if isinstance(text, str) and text.startswith("="):
        return text
    return None


def _extract_refs(formula: str, default_sheet: str) -> list[tuple[str, str]]:
    """Return ordered unique (sheet, a1) references found in a formula string."""
    out: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    # Drop the leading '=' and any quoted string literals to avoid false matches.
    body = formula[1:] if formula.startswith("=") else formula
    for match in _REF.finditer(body):
        sheet = match.group("qsheet") or match.group("sheet") or default_sheet
        a1 = match.group("a1").replace("$", "")
        key = (sheet, a1)
        if key not in seen:
            seen.add(key)
            out.append(key)
    return out


def _range_values(ws: Any, resolver: Any, sheet: str, a1: str) -> list[Any] | None:
    min_col, min_row, max_col, max_row = range_boundaries(a1)
    count = (max_col - min_col + 1) * (max_row - min_row + 1)
    if count > _RANGE_VALUE_CAP:
        return None
    values: list[Any] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            coord = f"{get_column_letter(c)}{r}"
            values.append(cell_value_for_json(resolver.resolve(sheet, coord, ws.cell(r, c).value)))
    return values


def _build_node(session: WorkbookSession, sheet: str, a1: str, remaining_depth: int) -> dict[str, Any]:
    resolver = get_value_resolver(session)
    node: dict[str, Any] = {
        "ref": f"{sheet}!{a1}",
        "sheet": sheet,
        "cell": a1,
        "is_range": ":" in a1,
    }
    if sheet not in session.workbook.sheetnames:
        node["error"] = "sheet_not_found"
        return node
    ws = session.workbook[sheet]

    if node["is_range"]:
        node["values"] = _range_values(ws, resolver, sheet, a1)
        # Recurse into the range's cells' precedents (bounded) so cross-sheet chains resolve.
        if remaining_depth > 0:
            min_col, min_row, max_col, max_row = range_boundaries(a1)
            if (max_col - min_col + 1) * (max_row - min_row + 1) <= _RANGE_VALUE_CAP:
                precedents: list[dict[str, Any]] = []
                seen: set[tuple[str, str]] = set()
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        formula = _formula_text(ws.cell(r, c).value)
                        if not formula:
                            continue
                        for psheet, pa1 in _extract_refs(formula, sheet):
                            if (psheet, pa1) in seen:
                                continue
                            seen.add((psheet, pa1))
                            precedents.append(_build_node(session, psheet, pa1, remaining_depth - 1))
                if precedents:
                    node["precedents"] = precedents
        return node

    raw = ws[a1].value
    node["value"] = cell_value_for_json(resolver.resolve(sheet, a1, raw))
    formula = _formula_text(raw)
    node["formula"] = formula
    if formula and remaining_depth > 0:
        node["precedents"] = [
            _build_node(session, psheet, pa1, remaining_depth - 1)
            for psheet, pa1 in _extract_refs(formula, sheet)
        ]
    return node


def trace_cell(session: WorkbookSession, sheet: str, cell: str, depth: int) -> SpreadsheetTraceResponse:
    """Trace a cell's formula lineage: its precedents (and theirs, up to `depth`)."""
    if sheet not in session.workbook.sheetnames:
        raise PolicyError("sheet_not_found", "Sheet does not exist.")
    depth = max(0, min(depth, MAX_TRACE_DEPTH))
    cell = cell.replace("$", "").upper()
    target = _build_node(session, sheet, cell, depth)
    ref = source_ref(session.path, sheet, cell)
    return SpreadsheetTraceResponse(
        ok=True,
        session_id=session.session_id,
        source_refs=[ref],
        target=target,
        depth=depth,
    )
