from __future__ import annotations

from openpyxl.workbook.defined_name import DefinedNameDict

from excel_mcp.best_source import rank_sources
from excel_mcp.regions import detect_regions
from excel_mcp.schemas import SheetInfo, SpreadsheetDescribeResponse
from excel_mcp.session import WorkbookSession
from excel_mcp.utils import bounds_to_a1, source_ref


def describe_workbook(session: WorkbookSession, detail: str = "compact") -> SpreadsheetDescribeResponse:
    workbook = session.workbook
    full_sheets: list[SheetInfo] = []
    all_regions = []

    for ws in workbook.worksheets:
        bounds = bounds_to_a1(ws.min_row, ws.min_column, ws.max_row, ws.max_column)
        regions = detect_regions(session.path, ws)
        all_regions.extend(regions)
        formula_count = sum(
            1
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        full_sheets.append(
            SheetInfo(
                name=ws.title,
                bounds=bounds,
                hidden=ws.sheet_state != "visible",
                merged_ranges_count=len(ws.merged_cells.ranges),
                merged_ranges_sample=[str(item) for item in list(ws.merged_cells.ranges)[:10]],
                formula_count=formula_count,
                named_ranges=_named_ranges_for_sheet(workbook.defined_names, ws.title),
                excel_tables=list(ws.tables.keys()),
                regions=regions,
                source_ref=source_ref(session.path, ws.title, bounds),
            )
        )

    session.regions = all_regions
    # Rank on the full (uncompacted) regions so sample-row text is available to the
    # ranker, then compact for the response — compacting first would blind
    # rank_sources to matrix-layout summary regions (see best_source._looks_like_summary).
    best_source = rank_sources(full_sheets)[:3]
    sheets = [_compact_sheet(sheet) for sheet in full_sheets] if detail == "compact" else full_sheets

    response = SpreadsheetDescribeResponse(
        ok=True,
        session_id=session.session_id,
        source_refs=[sheet.source_ref for sheet in sheets],
        file_name=session.path.name,
        sheet_count=len(sheets),
        sheets=sheets,
    )
    response.best_source = best_source
    return response


def _compact_sheet(sheet: SheetInfo) -> SheetInfo:
    copied = sheet.model_copy(deep=True)
    copied.regions = [_compact_region(region) for region in copied.regions]
    return copied


def _compact_region(region):
    copied = region.model_copy(deep=True)
    copied.sample_rows = []
    return copied


def _named_ranges_for_sheet(defined_names: DefinedNameDict, sheet: str) -> list[str]:
    matches: list[str] = []
    for name, defined_name in defined_names.items():
        text = defined_name.attr_text or ""
        if f"'{sheet}'!" in text or f"{sheet}!" in text:
            matches.append(name)
    return matches
