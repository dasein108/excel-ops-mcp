from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter


def source_ref(path: Path, sheet: str, bounds: str) -> str:
    safe_sheet = sheet.replace("'", "''")
    return f"{path.name}#'{safe_sheet}'!{bounds}"


def bounds_to_a1(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def safe_identifier(name: str, fallback: str) -> str:
    text = re.sub(r"\W+", "_", str(name).strip()).strip("_").lower()
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"c_{text}"
    return text


def dedupe_names(names: list[str]) -> list[str]:
    seen: Counter[str] = Counter()
    out: list[str] = []
    for name in names:
        base = safe_identifier(name, "column")
        seen[base] += 1
        out.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return out


def cell_value_for_json(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value

