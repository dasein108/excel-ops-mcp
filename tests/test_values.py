from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from excel_mcp.config import ExcelMcpConfig
from excel_mcp.tools import ExcelMcpTools
from excel_mcp.values import ValueResolver, is_formula

EXAMPLES = Path(__file__).resolve().parents[1] / "examples"
REPO_ROOT = Path(__file__).resolve().parents[1]


def _recompute_available() -> bool:
    try:
        import formulas  # noqa: F401

        return True
    except Exception:
        return False


class CachedValueResolutionTests(unittest.TestCase):
    """The saas.xlsx model was saved by a spreadsheet app, so formula cells carry
    cached values. These must surface as numbers, not formula strings."""

    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.tools = ExcelMcpTools(
            ExcelMcpConfig(allowed_roots=[REPO_ROOT], cache_dir=Path(self.tmp.name) / ".cache")
        )
        response = self.tools.spreadsheet_open({"path": str(EXAMPLES / "saas.xlsx")})
        self.assertTrue(response["ok"], response)
        self.session_id = response["session_id"]

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_read_range_returns_computed_numbers_not_formulas(self) -> None:
        response = self.tools.spreadsheet_read_range(
            {
                "session_id": self.session_id,
                "sheet": "Revenue Model",
                "range": "B23:D23",  # Total Revenue = B11+B16+B20 per month
                "include": ["values"],
            }
        )
        self.assertTrue(response["ok"], response)
        values = [cell["value"] for row in response["cells"] for cell in row]
        self.assertEqual(values[0], 87680)
        for value in values:
            self.assertIsInstance(value, (int, float))

    def test_read_range_formulas_include_still_returns_formula_text(self) -> None:
        response = self.tools.spreadsheet_read_range(
            {
                "session_id": self.session_id,
                "sheet": "Revenue Model",
                "range": "B23:B23",
                "include": ["values", "formulas"],
            }
        )
        cell = response["cells"][0][0]
        self.assertEqual(cell["value"], 87680)
        self.assertEqual(cell["formula"], "=B11+B16+B20")

    def test_array_formula_cell_resolves_to_value_not_object(self) -> None:
        # Dashboard!B15 (break-even month) is an ArrayFormula; must surface as a number.
        response = self.tools.spreadsheet_read_range(
            {
                "session_id": self.session_id,
                "sheet": "Dashboard",
                "range": "B15:B15",
                "include": ["values"],
            }
        )
        value = response["cells"][0][0]["value"]
        self.assertEqual(value, 9)
        self.assertNotIn("ArrayFormula", str(value))

    def test_query_aggregates_computed_formula_column(self) -> None:
        try:
            import duckdb  # noqa: F401
        except Exception:
            self.skipTest("duckdb is not installed")
        # Dashboard is a matrix region: line_item + year_1..year_5. Gross Profit Y1.
        response = self.tools.spreadsheet_query(
            {
                "session_id": self.session_id,
                "sql": "select line_item, year_1 from dashboard_table_2 "
                "where line_item like '%Gross Profit%'",
            }
        )
        self.assertTrue(response["ok"], response)
        y1 = float(response["rows"][0]["year_1"])
        self.assertAlmostEqual(y1, 1119393.6, places=1)


class RecomputeFallbackTests(unittest.TestCase):
    """openpyxl-authored workbooks have no cached values, exercising the fallback."""

    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "fresh.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(self.path)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_literal_cells_pass_through(self) -> None:
        resolver = ValueResolver(self.path)
        self.assertEqual(resolver.resolve("Sheet1", "A1", 10), 10)
        self.assertFalse(is_formula(10))
        self.assertTrue(is_formula("=SUM(A1:A2)"))

    def test_uncached_formula_degrades_or_recomputes(self) -> None:
        resolver = ValueResolver(self.path)
        result = resolver.resolve("Sheet1", "A3", "=SUM(A1:A2)")
        if _recompute_available():
            self.assertEqual(result, 30)
        else:
            self.assertIsNone(result)
            self.assertIn("computed_value_unavailable", resolver.warnings)


if __name__ == "__main__":
    unittest.main()
