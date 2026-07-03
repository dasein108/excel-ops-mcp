from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from excel_mcp.config import ExcelMcpConfig
from excel_mcp.policy import PathPolicy, PolicyError, SqlPolicy
from excel_mcp.tools import ExcelMcpTools


def make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Date", "Vendor", "Category", "Amount"])
    ws.append(["2026-01-01", "Acme", "Software", 100])
    ws.append(["2026-01-03", "Beta", "Marketing", 250])
    ws.append(["2026-01-10", "Acme", "Software", 150])

    ws["F1"] = "Metric"
    ws["G1"] = "Value"
    ws["F2"] = "Total"
    ws["G2"] = "=SUM(D2:D4)"

    summary = wb.create_sheet("Summary")
    summary.append(["Category", "Budget"])
    summary.append(["Software", 500])
    summary.append(["Marketing", 400])

    wb.save(path)


class ReadOnlyExcelMcpTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.workbook = self.root / "ops.xlsx"
        make_workbook(self.workbook)
        self.tools = ExcelMcpTools(ExcelMcpConfig(allowed_roots=[self.root], cache_dir=self.root / ".cache"))

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def open_session(self) -> str:
        response = self.tools.spreadsheet_open({"path": str(self.workbook)})
        self.assertTrue(response["ok"], response)
        return response["session_id"]

    def test_path_policy_blocks_outside_root(self) -> None:
        policy = PathPolicy(ExcelMcpConfig(allowed_roots=[self.root]))
        with self.assertRaises(PolicyError):
            policy.validate_input_file("/tmp/outside.xlsx")

    def test_sql_policy_blocks_mutation(self) -> None:
        policy = SqlPolicy()
        with self.assertRaises(PolicyError):
            policy.validate_readonly("delete from Transactions")

    def test_open_reuses_session_for_unchanged_file(self) -> None:
        first = self.tools.spreadsheet_open({"path": str(self.workbook)})
        second = self.tools.spreadsheet_open({"path": str(self.workbook)})
        self.assertTrue(first["ok"])
        self.assertTrue(second["ok"])
        self.assertEqual(first["session_id"], second["session_id"])
        self.assertEqual(second["telemetry"]["cache"], "hit")

    def test_describe_detects_regions_and_formulas(self) -> None:
        session_id = self.open_session()
        response = self.tools.spreadsheet_describe({"session_id": session_id})
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["sheet_count"], 2)
        transactions = next(sheet for sheet in response["sheets"] if sheet["name"] == "Transactions")
        self.assertGreaterEqual(transactions["formula_count"], 1)
        self.assertGreaterEqual(len(transactions["regions"]), 2)
        self.assertTrue(any(region["region_kind"] in {"ledger", "table"} for region in transactions["regions"]))

    def test_read_range_returns_cells(self) -> None:
        session_id = self.open_session()
        response = self.tools.spreadsheet_read_range(
            {"session_id": session_id, "sheet": "Transactions", "range": "A1:B2", "include": ["values"]}
        )
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["cells"][0][0]["value"], "Date")
        self.assertEqual(response["cells"][1][1]["value"], "Acme")

    def test_repeated_small_range_read_warns(self) -> None:
        session_id = self.open_session()
        payload = {"session_id": session_id, "sheet": "Transactions", "range": "A1:B2", "include": ["values"]}

        first = self.tools.spreadsheet_read_range(dict(payload))
        second = self.tools.spreadsheet_read_range(dict(payload))
        self.assertEqual(first["warnings"], [])
        self.assertEqual(second["warnings"], [])

        third = self.tools.spreadsheet_read_range(dict(payload))
        self.assertTrue(third["ok"], third)
        self.assertTrue(any("read 3 times" in warning for warning in third["warnings"]), third["warnings"])

        # A different range does not inherit the repeated-read warning.
        other = self.tools.spreadsheet_read_range(
            {"session_id": session_id, "sheet": "Transactions", "range": "C1:D2", "include": ["values"]}
        )
        self.assertEqual(other["warnings"], [])

    def test_query_sums_amounts_when_duckdb_installed(self) -> None:
        try:
            import duckdb  # noqa: F401
        except Exception:
            self.skipTest("duckdb is not installed")

        session_id = self.open_session()
        describe = self.tools.spreadsheet_describe({"session_id": session_id})
        ledger = next(region for sheet in describe["sheets"] for region in sheet["regions"] if region["region_kind"] == "ledger")
        response = self.tools.spreadsheet_query(
            {"session_id": session_id, "sql": f'select vendor, sum(amount) as total from "{ledger["table_name"]}" group by vendor order by vendor'}
        )
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["rows"][0]["vendor"], "Acme")
        self.assertEqual(response["rows"][0]["total"], 250)

    def test_query_truncates_with_telemetry(self) -> None:
        try:
            import duckdb  # noqa: F401
        except Exception:
            self.skipTest("duckdb is not installed")

        tools = ExcelMcpTools(ExcelMcpConfig(allowed_roots=[self.root], cache_dir=self.root / ".cache", query_row_limit=1))
        open_response = tools.spreadsheet_open({"path": str(self.workbook)})
        session_id = open_response["session_id"]
        describe = tools.spreadsheet_describe({"session_id": session_id})
        ledger = next(region for sheet in describe["sheets"] for region in sheet["regions"] if region["region_kind"] == "ledger")
        response = tools.spreadsheet_query(
            {"session_id": session_id, "sql": f'select vendor, amount from "{ledger["table_name"]}"'}
        )
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["row_count"], 1)
        self.assertTrue(response["telemetry"]["truncated"], response["telemetry"])


if __name__ == "__main__":
    unittest.main()

