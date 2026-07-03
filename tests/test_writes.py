from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_mcp.config import ExcelMcpConfig
from excel_mcp.tools import ExcelMcpTools


def make_write_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ops"
    ws.append(["Date", "Vendor", "Category", "Amount", "Review"])
    ws.append(["2026-01-01", "Acme", "Software", 100, None])
    ws.append(["2026-01-02", "Beta", "Marketing", 250, None])
    ws["D4"] = "=SUM(D2:D3)"
    ws["E4"] = "Total"
    wb.save(path)


class SafeWriteTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.workbook = self.root / "ops.xlsx"
        make_write_workbook(self.workbook)
        self.tools = ExcelMcpTools(ExcelMcpConfig(allowed_roots=[self.root], cache_dir=self.root / ".cache"))
        opened = self.tools.spreadsheet_open({"path": str(self.workbook)})
        self.assertTrue(opened["ok"], opened)
        self.session_id = opened["session_id"]

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_write_dry_run_stages_without_mutating_source(self) -> None:
        before = load_workbook(self.workbook, data_only=False)["Ops"]["E2"].value

        response = self.tools.spreadsheet_write(
            {
                "session_id": self.session_id,
                "dry_run": True,
                "operations": [{"type": "set_values", "sheet": "Ops", "start": "E2", "values": [["reviewed"]]}],
            }
        )

        self.assertTrue(response["ok"], response)
        self.assertIsNotNone(response["staged_id"])
        self.assertEqual(response["accepted_operations"], 1)
        self.assertEqual(response["touched_ranges"], ["Ops!E2:E2"])
        self.assertEqual(response["changes"][0]["before"], before)
        self.assertEqual(response["changes"][0]["after"], "reviewed")

        after = load_workbook(self.workbook, data_only=False)["Ops"]["E2"].value
        self.assertEqual(after, before)

    def test_diff_returns_staged_changes(self) -> None:
        write = self.tools.spreadsheet_write(
            {
                "session_id": self.session_id,
                "operations": [{"type": "set_formula", "sheet": "Ops", "cell": "E5", "formula": "=D4"}],
            }
        )

        diff = self.tools.spreadsheet_diff({"session_id": self.session_id, "staged_id": write["staged_id"]})
        self.assertTrue(diff["ok"], diff)
        self.assertEqual(diff["changed_ranges"], ["Ops!E5:E5"])
        self.assertEqual(diff["changes"][0]["after"], "=D4")

    def test_commit_saves_to_new_file_and_preserves_source(self) -> None:
        write = self.tools.spreadsheet_write(
            {
                "session_id": self.session_id,
                "operations": [
                    {"type": "set_values", "sheet": "Ops", "start": "E2", "values": [["reviewed"]]},
                    {"type": "set_formula", "sheet": "Ops", "cell": "E5", "formula": "=D4"},
                ],
            }
        )
        output = self.root / "ops.reviewed.xlsx"

        commit = self.tools.spreadsheet_commit(
            {
                "session_id": self.session_id,
                "staged_id": write["staged_id"],
                "output_path": str(output),
            }
        )

        self.assertTrue(commit["ok"], commit)
        self.assertEqual(Path(commit["output_path"]).resolve(), output.resolve())

        source_ws = load_workbook(self.workbook, data_only=False)["Ops"]
        output_ws = load_workbook(output, data_only=False)["Ops"]
        self.assertIsNone(source_ws["E2"].value)
        self.assertEqual(source_ws["D4"].value, "=SUM(D2:D3)")
        self.assertEqual(output_ws["E2"].value, "reviewed")
        self.assertEqual(output_ws["E5"].value, "=D4")
        self.assertEqual(output_ws["D4"].value, "=SUM(D2:D3)")

    def test_commit_refuses_overwrite_by_default(self) -> None:
        write = self.tools.spreadsheet_write(
            {
                "session_id": self.session_id,
                "operations": [{"type": "set_values", "sheet": "Ops", "start": "E2", "values": [["reviewed"]]}],
            }
        )

        commit = self.tools.spreadsheet_commit(
            {
                "session_id": self.session_id,
                "staged_id": write["staged_id"],
                "output_path": str(self.workbook),
            }
        )

        self.assertFalse(commit["ok"])
        self.assertEqual(commit["error"]["code"], "output_exists")

    def test_clear_range_warns_when_touching_formula(self) -> None:
        response = self.tools.spreadsheet_write(
            {
                "session_id": self.session_id,
                "operations": [{"type": "clear_range", "sheet": "Ops", "range": "D4:E4"}],
            }
        )

        self.assertTrue(response["ok"], response)
        self.assertTrue(any("formula cell" in warning for warning in response["warnings"]))


if __name__ == "__main__":
    unittest.main()
