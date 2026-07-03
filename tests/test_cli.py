from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SAAS = ROOT / "examples" / "saas.xlsx"


def make_cli_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ops"
    ws.append(["Vendor", "Amount", "Status"])
    ws.append(["Acme", 100, None])
    ws.append(["Beta", 200, None])
    ws["B4"] = "=SUM(B2:B3)"
    wb.save(path)


class ExcelOpsCliTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.cache = self.root / "cache"

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def run_cli(self, *args: str, expect: int = 0) -> dict:
        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "excel_mcp.cli",
                *args,
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        )
        self.assertEqual(result.returncode, expect, result.stderr + result.stdout)
        return json.loads(result.stdout)

    def common(self) -> list[str]:
        return ["--allowed-root", str(ROOT), "--cache-dir", str(self.cache)]

    def test_stateless_describe_example(self) -> None:
        response = self.run_cli("describe", str(SAAS), *self.common())
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["sheet_count"], 5)
        self.assertEqual(response["sheets"][0]["name"], "Assumptions")

    def test_sheets_lists_workbook_sheets_compactly(self) -> None:
        response = self.run_cli("sheets", str(SAAS), *self.common())
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["sheet_count"], 5)
        self.assertEqual(response["sheets"][0]["name"], "Assumptions")
        self.assertEqual(response["sheets"][0]["region_count"], 12)

    def test_tables_lists_detected_regions_for_one_sheet(self) -> None:
        response = self.run_cli(
            "tables",
            str(SAAS),
            "--sheet",
            "Revenue Model",
            *self.common(),
        )
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["table_count"], 6)
        self.assertEqual(response["tables"][0]["table_name"], "revenue_model_table_1")
        self.assertIn("line_item", response["tables"][0]["columns"])

    def test_stateless_query_example(self) -> None:
        response = self.run_cli(
            "query",
            str(SAAS),
            "--sql",
            'select count(*) as n from "revenue_model_table_1"',
            *self.common(),
        )
        self.assertTrue(response["ok"], response)
        self.assertEqual(response["rows"][0]["n"], 1)

    def test_open_then_describe_by_session_across_processes(self) -> None:
        opened = self.run_cli("open", str(SAAS), *self.common())
        self.assertTrue(opened["ok"], opened)

        described = self.run_cli("describe", "--session", opened["session_id"], *self.common())
        self.assertTrue(described["ok"], described)
        self.assertEqual(described["sheets"][0]["name"], "Assumptions")

    def test_write_diff_commit_across_processes(self) -> None:
        workbook = self.root / "ops.xlsx"
        output = self.root / "ops.updated.xlsx"
        ops_path = self.root / "ops.json"
        make_cli_workbook(workbook)
        ops_path.write_text(
            json.dumps([{"type": "set_values", "sheet": "Ops", "start": "C2", "values": [["reviewed"]]}]),
            encoding="utf-8",
        )

        common = ["--allowed-root", str(self.root), "--cache-dir", str(self.cache)]
        opened = self.run_cli("open", str(workbook), *common)
        written = self.run_cli("write", "--session", opened["session_id"], "--ops", str(ops_path), *common)
        self.assertTrue(written["ok"], written)
        self.assertEqual(written["touched_ranges"], ["Ops!C2:C2"])

        diffed = self.run_cli("diff", "--session", opened["session_id"], "--staged", written["staged_id"], *common)
        self.assertTrue(diffed["ok"], diffed)
        self.assertEqual(diffed["changes"][0]["after"], "reviewed")

        committed = self.run_cli(
            "commit",
            "--session",
            opened["session_id"],
            "--staged",
            written["staged_id"],
            "--output",
            str(output),
            *common,
        )
        self.assertTrue(committed["ok"], committed)
        self.assertIsNone(load_workbook(workbook)["Ops"]["C2"].value)
        self.assertEqual(load_workbook(output)["Ops"]["C2"].value, "reviewed")

    def test_invalid_sql_returns_exit_1_json_error(self) -> None:
        response = self.run_cli("query", str(SAAS), "--sql", 'delete from "revenue_model_table_1"', *self.common(), expect=1)
        self.assertFalse(response["ok"])
        self.assertEqual(response["error"]["code"], "mutating_sql_rejected")


if __name__ == "__main__":
    unittest.main()
